"""Upload pipeline tests.

These call the worker function `process_upload` directly so we exercise the
ingest code path end-to-end without going through Redis. The shape is the
same; only the queue round-trip is skipped.
"""
from __future__ import annotations

import shutil
import uuid
from pathlib import Path

import openpyxl
import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.db import engine
from app.main import app
from app.models.buyer_product_stat import BuyerProductStat
from app.models.order import Order
from app.models.upload_batch import UploadBatch
from app.models.user import User
from app.models.user_settings import UserSettings
from app.services.auth import register_user
from app.services.buyer_key import derive_buyer_key
from app.workers.db import SyncSessionLocal
from app.workers.upload import process_upload

FIXTURE_DIR = Path(__file__).parent / "fixtures"
SAMPLE_FIXTURE = FIXTURE_DIR / "sample_orders.xlsx"
TMP_DIR = Path("/tmp/uploads")
TEST_PW = "letter1-letter2"  # pragma: allowlist secret


@pytest_asyncio.fixture
async def client() -> AsyncClient:
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c
    await engine.dispose()


def _email() -> str:
    return f"test-{uuid.uuid4().hex[:10]}@example.com"


async def _make_user_and_batch(filename: str = "sample.xlsx") -> tuple[User, UploadBatch, Path]:
    """Create a user via the auth service, drop a batch row, and stage a
    copy of the fixture under /tmp/uploads/<batch_id>.xlsx (the path the
    worker expects)."""
    from app.core.db import SessionLocal as AsyncSessionLocal

    TMP_DIR.mkdir(parents=True, exist_ok=True)

    async with AsyncSessionLocal() as session:
        user = await register_user(session, _email(), TEST_PW)

    batch_id = uuid.uuid4()
    file_path = TMP_DIR / f"{batch_id}.xlsx"
    shutil.copyfile(SAMPLE_FIXTURE, file_path)

    with SyncSessionLocal() as sync_session:
        batch = UploadBatch(
            id=batch_id,
            user_id=user.id,
            filename=filename,
            file_size_bytes=file_path.stat().st_size,
            status="processing",
        )
        sync_session.add(batch)
        sync_session.commit()
        sync_session.refresh(batch)

    return user, batch, file_path


# ---------- happy path ----------

async def test_happy_path_10_rows(client: AsyncClient) -> None:
    user, batch, path = await _make_user_and_batch()
    process_upload(str(batch.id), str(user.id), str(path))

    with SyncSessionLocal() as session:
        refreshed = session.get(UploadBatch, batch.id)
        assert refreshed is not None
        assert refreshed.status == "completed"
        assert refreshed.total_rows == 10
        assert refreshed.new_rows == 10
        assert refreshed.updated_rows == 0
        assert refreshed.duplicate_rows == 0
        assert refreshed.error_rows == 0

        order_count = session.scalar(
            select(_count_user_orders(user.id))
        )
        assert order_count == 10


async def test_reupload_is_all_duplicates(client: AsyncClient) -> None:
    user, batch, path = await _make_user_and_batch()
    process_upload(str(batch.id), str(user.id), str(path))

    # Second upload of the same fixture, fresh batch row + temp file.
    second_batch_id = uuid.uuid4()
    second_path = TMP_DIR / f"{second_batch_id}.xlsx"
    shutil.copyfile(SAMPLE_FIXTURE, second_path)
    with SyncSessionLocal() as session:
        session.add(
            UploadBatch(
                id=second_batch_id,
                user_id=user.id,
                filename="sample-2.xlsx",
                file_size_bytes=second_path.stat().st_size,
                status="processing",
            )
        )
        session.commit()

    process_upload(str(second_batch_id), str(user.id), str(second_path))

    with SyncSessionLocal() as session:
        refreshed = session.get(UploadBatch, second_batch_id)
        assert refreshed is not None
        assert refreshed.status == "completed"
        # Every row maps to an existing (user_id, order_id) → 0 new, 10
        # skipped (reported in updated_rows for historical compatibility).
        assert refreshed.new_rows == 0
        assert refreshed.updated_rows == 10
        # orders table size stays at 10
        order_count = session.scalar(select(_count_user_orders(user.id)))
        assert order_count == 10


async def test_reupload_does_not_modify_existing_row(client: AsyncClient) -> None:
    """Skipping duplicates means the existing row's columns must NOT change,
    even when the new file disagrees with the DB."""
    user, batch, path = await _make_user_and_batch()
    process_upload(str(batch.id), str(user.id), str(path))

    # Snapshot one order's tracking number and item_price before the
    # second upload.
    with SyncSessionLocal() as session:
        original = session.scalar(
            select(Order).where(Order.user_id == user.id, Order.order_id == "O1")
        )
        assert original is not None
        original_tracking = original.tracking_number
        original_price = original.item_price

    # Build a second fixture where order O1 has a different tracking
    # number and price. After re-upload the existing row should still
    # carry the original values — proof that duplicates are skipped, not
    # upserted.
    second_batch_id = uuid.uuid4()
    second_path = TMP_DIR / f"{second_batch_id}.xlsx"
    shutil.copyfile(SAMPLE_FIXTURE, second_path)
    wb = openpyxl.load_workbook(second_path)
    ws = wb["配送信息"]
    headers = [c.value for c in ws[1]]
    tracking_col = headers.index("追踪号码") + 1
    price_col = headers.index("商品售价") + 1
    order_col = headers.index("订单编号") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=order_col).value == "O1":
            ws.cell(row=row, column=tracking_col).value = "TRK-CHANGED"
            ws.cell(row=row, column=price_col).value = 999.99
            break
    wb.save(second_path)

    with SyncSessionLocal() as session:
        session.add(
            UploadBatch(
                id=second_batch_id,
                user_id=user.id,
                filename="sample-mutated.xlsx",
                file_size_bytes=second_path.stat().st_size,
                status="processing",
            )
        )
        session.commit()

    process_upload(str(second_batch_id), str(user.id), str(second_path))

    with SyncSessionLocal() as session:
        refreshed = session.get(UploadBatch, second_batch_id)
        assert refreshed is not None
        assert refreshed.new_rows == 0
        assert refreshed.updated_rows == 10  # all 10 rows skipped

        unchanged = session.scalar(
            select(Order).where(Order.user_id == user.id, Order.order_id == "O1")
        )
        assert unchanged is not None
        assert unchanged.tracking_number == original_tracking
        assert unchanged.item_price == original_price


async def test_buyer_key_falls_back_to_address(client: AsyncClient) -> None:
    user, batch, path = await _make_user_and_batch()
    process_upload(str(batch.id), str(user.id), str(path))

    expected = derive_buyer_key(None, "p3:CA", "CA", "ON", "Toronto")
    with SyncSessionLocal() as session:
        rows = (
            session.execute(
                select(Order)
                .where(Order.user_id == user.id)
                .where(Order.order_id.in_(["O5", "O6"]))
            )
            .scalars()
            .all()
        )
        assert {r.buyer_key for r in rows} == {expected}


async def test_buyer_product_stats_counts(client: AsyncClient) -> None:
    user, batch, path = await _make_user_and_batch()
    process_upload(str(batch.id), str(user.id), str(path))

    alice_key = derive_buyer_key("alice@example.com", "p3:US", "US", "TX", "Austin")
    toronto_key = derive_buyer_key(None, "p3:CA", "CA", "ON", "Toronto")

    with SyncSessionLocal() as session:
        # Alice + p3:US + A1 → 3 orders, across all three grains
        for grain, value in (("asin", "A1"), ("spu", "S-WIDGET"),
                             ("product_name", "Widget A1")):
            stat = session.execute(
                select(BuyerProductStat).where(
                    BuyerProductStat.user_id == user.id,
                    BuyerProductStat.shop_site == "p3:US",
                    BuyerProductStat.buyer_key == alice_key,
                    BuyerProductStat.grain == grain,
                    BuyerProductStat.group_value == value,
                )
            ).scalar_one()
            assert stat.order_count == 3

        # Toronto address buyer + p3:CA + A1 → 2 orders
        stat = session.execute(
            select(BuyerProductStat).where(
                BuyerProductStat.user_id == user.id,
                BuyerProductStat.shop_site == "p3:CA",
                BuyerProductStat.buyer_key == toronto_key,
                BuyerProductStat.grain == "asin",
                BuyerProductStat.group_value == "A1",
            )
        ).scalar_one()
        assert stat.order_count == 2


async def test_first_upload_hook_sets_active_shop(client: AsyncClient) -> None:
    user, batch, path = await _make_user_and_batch()
    process_upload(str(batch.id), str(user.id), str(path))

    with SyncSessionLocal() as session:
        settings_row = session.get(UserSettings, user.id)
        assert settings_row is not None
        # p3:US has the most rows in the fixture (Alice ×4 + Bob + Dave = 6),
        # so the hook should pick it.
        assert settings_row.active_shop_site == "p3:US"


# ---------- bad-shape paths ----------

async def test_missing_required_column_fails_batch(
    client: AsyncClient, tmp_path: Path
) -> None:
    """Drop the `订单编号` column and verify the batch ends `failed`."""
    user, batch, _path = await _make_user_and_batch()

    # Build a corrupt fixture by removing 订单编号.
    src = openpyxl.load_workbook(SAMPLE_FIXTURE)
    sheet = src.active
    header_row = [c.value for c in sheet[1]]
    drop_idx = header_row.index("订单编号") + 1
    sheet.delete_cols(drop_idx)
    bad_path = tmp_path / "bad.xlsx"
    src.save(bad_path)

    process_upload(str(batch.id), str(user.id), str(bad_path))

    with SyncSessionLocal() as session:
        refreshed = session.get(UploadBatch, batch.id)
        assert refreshed is not None
        assert refreshed.status == "failed"
        assert refreshed.error_detail is not None
        assert "missing" in refreshed.error_detail.get("reason", "").lower()
        # No orders should have been inserted.
        order_count = session.scalar(select(_count_user_orders(user.id)))
        assert order_count == 0


async def test_empty_order_id_counts_as_error_row(
    client: AsyncClient, tmp_path: Path
) -> None:
    """One row's 订单编号 set to empty string; the rest should still ingest."""
    user, batch, _orig = await _make_user_and_batch()

    src = openpyxl.load_workbook(SAMPLE_FIXTURE)
    sheet = src.active
    header_row = [c.value for c in sheet[1]]
    col_idx = header_row.index("订单编号") + 1
    # Row 2 (first data row in openpyxl 1-indexed) — wipe its order_id.
    # Note: cell(row, column, value=...) treats value=None as "no value
    # supplied" (a no-op), so we must set .value explicitly to clear.
    sheet.cell(row=2, column=col_idx).value = None
    bad_path = tmp_path / "missing-order-id.xlsx"
    src.save(bad_path)

    process_upload(str(batch.id), str(user.id), str(bad_path))

    with SyncSessionLocal() as session:
        refreshed = session.get(UploadBatch, batch.id)
        assert refreshed is not None
        assert refreshed.status == "completed"
        assert refreshed.error_rows == 1
        assert refreshed.new_rows == 9


def _count_user_orders(user_id):
    from sqlalchemy import func
    return func.count(Order.id).filter(Order.user_id == user_id)
